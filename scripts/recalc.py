"""
scripts/recalc.py
Verifies GPBullhound_LiveData.xlsx has zero formula error cells.
Scans every cell across all sheets and reports #DIV/0!, #REF!, #N/A,
#NAME?, #VALUE!, #NULL!, #NUM! occurrences.
"""

import os
import sys
from openpyxl import load_workbook

EXCEL_ERRORS = {"#DIV/0!", "#REF!", "#N/A", "#NAME?", "#VALUE!", "#NULL!", "#NUM!"}

WORKBOOK_PATH = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "GPBullhound_LiveData.xlsx"
)


def verify_workbook(path: str) -> int:
    """
    Load workbook and scan all cells.
    Returns the total number of formula-error cells found.
    """
    if not os.path.exists(path):
        print(f"ERROR: Workbook not found at {path}")
        print("Run:  python scripts/generate_excel.py   first.")
        return 1

    print(f"Loading {path} …")
    wb = load_workbook(path, data_only=False)   # keep formulas, not cached values

    total_cells   = 0
    error_cells   = 0
    formula_cells = 0
    errors_found  = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_errors = 0
        for row in ws.iter_rows():
            for cell in row:
                total_cells += 1
                val = cell.value
                if val is None:
                    continue
                val_str = str(val).strip()
                # Count formula cells (start with '=')
                if val_str.startswith("="):
                    formula_cells += 1
                # Check for Excel error strings in static values
                if val_str.upper() in {e.upper() for e in EXCEL_ERRORS}:
                    error_cells += 1
                    sheet_errors += 1
                    errors_found.append(
                        f"  [{sheet_name}!{cell.coordinate}] → {val_str}"
                    )
        if sheet_errors:
            print(f"  {sheet_name}: {sheet_errors} error(s)")

    print()
    print("── Recalc Summary ──────────────────────────────────────────")
    print(f"  Sheets scanned  : {len(wb.sheetnames)}")
    print(f"  Total cells     : {total_cells:,}")
    print(f"  Formula cells   : {formula_cells:,}")
    print(f"  Error cells     : {error_cells}")
    print("────────────────────────────────────────────────────────────")

    if errors_found:
        print("\nERROR DETAILS:")
        for e in errors_found:
            print(e)
        print()
        print("RESULT: FAIL — formula errors detected.")
        return len(errors_found)

    print("\nRESULT: PASS — zero formula errors detected.")
    return 0


def check_sheet_names(wb):
    """Verify all 10 expected sheets are present."""
    expected = [
        "Portfolio", "Funds", "LPs", "ExitPipeline", "Pipeline",
        "MacroScenarios", "CapitalCalls", "Benchmarks", "WatchList", "Instructions"
    ]
    actual = wb.sheetnames
    missing = [s for s in expected if s not in actual]
    if missing:
        print(f"WARN: Missing sheets: {missing}")
    else:
        print(f"  Sheet names: OK ({len(actual)} sheets present)")
    return missing


def check_portfolio_headers(wb):
    """Verify Portfolio sheet has exact required headers."""
    required = [
        "name", "arr", "arr_gr", "nav", "nav_pct", "moic", "burn",
        "gm", "nrr", "ev_arr", "ev_arr_entry", "valuation",
        "sector", "stage", "geo", "status", "funds",
        "esg_e", "esg_s", "esg_g", "lever"
    ]
    ws = wb["Portfolio"]
    actual = [ws.cell(row=1, column=i+1).value for i in range(len(required))]
    mismatches = [(req, act) for req, act in zip(required, actual) if req != act]
    if mismatches:
        print(f"  Portfolio headers: MISMATCH → {mismatches}")
    else:
        print(f"  Portfolio headers: OK ({len(required)} columns match)")
    return mismatches


def check_formula_presence(wb):
    """Spot-check that key formula cells actually contain formulas."""
    checks = [
        ("Portfolio",       "E2",  "nav_pct formula"),
        ("Portfolio",       "J2",  "ev_arr formula"),
        ("LPs",             "J2",  "unfunded formula"),
        ("ExitPipeline",    "K2",  "pwav_proceeds formula"),
        ("Pipeline",        "G2",  "ev_arr formula"),
        ("CapitalCalls",    "E3",  "cumulative_calls formula"),
        ("CapitalCalls",    "F3",  "cumulative_dist formula"),
        ("CapitalCalls",    "G3",  "DPI formula"),
        ("WatchList",       "D4",  "Sanity VLOOKUP"),
        ("WatchList",       "E4",  "Sanity IF status"),
        ("WatchList",       "D5",  "Headway VLOOKUP"),
        ("WatchList",       "E5",  "Headway IF status"),
    ]
    failures = []
    for sheet_name, cell_addr, desc in checks:
        if sheet_name not in wb.sheetnames:
            failures.append(f"  Missing sheet: {sheet_name}")
            continue
        val = str(wb[sheet_name][cell_addr].value or "")
        if not val.startswith("="):
            failures.append(f"  [{sheet_name}!{cell_addr}] {desc}: expected formula, got {val!r}")
    if failures:
        print("  Formula presence: FAIL")
        for f in failures:
            print(f)
    else:
        print(f"  Formula presence: OK ({len(checks)} spot checks passed)")
    return failures


def main():
    if not os.path.exists(WORKBOOK_PATH):
        print(f"ERROR: {WORKBOOK_PATH} not found.")
        print("Generate it first:  python scripts/generate_excel.py")
        sys.exit(1)

    wb = load_workbook(WORKBOOK_PATH, data_only=False)

    print("── Structural Checks ───────────────────────────────────────")
    missing_sheets   = check_sheet_names(wb)
    header_mismatches= check_portfolio_headers(wb)
    formula_failures = check_formula_presence(wb)
    print()

    error_count = verify_workbook(WORKBOOK_PATH)

    structural_issues = len(missing_sheets) + len(header_mismatches) + len(formula_failures)
    total_issues = error_count + structural_issues

    if total_issues == 0:
        print("\n✓ All checks passed. GPBullhound_LiveData.xlsx is valid.")
        sys.exit(0)
    else:
        print(f"\n✗ {total_issues} issue(s) found. See details above.")
        sys.exit(1)


if __name__ == "__main__":
    main()
