"""
GPBullhound_LiveData.xlsx generator
Creates a 10-sheet VC fund dashboard workbook with dark theme styling.
"""

from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
import os

# ── Colours (no leading #) ───────────────────────────────────────────────────
C_ROW_ODD    = "0D1014"
C_ROW_EVEN   = "0F1318"
C_HEADER_BG  = "131820"
C_HEADER_TXT = "7A776F"
C_INPUT_TXT  = "4A8BD4"
C_FORMULA_TXT= "2ECC8A"
C_BREACH_BG  = "E05555"
C_BORDER_HDR = "E8622A"
C_WHITE      = "FFFFFF"

# Tab colours per sheet
TAB_COLORS = [
    "4A8BD4",  # Portfolio
    "E8622A",  # Funds
    "2ECC8A",  # LPs
    "E05555",  # ExitPipeline
    "7A776F",  # Pipeline
    "8B4ACA",  # MacroScenarios
    "4A6BD4",  # CapitalCalls
    "D4A44A",  # Benchmarks
    "D44A4A",  # WatchList
    "4AD4A4",  # Instructions
]

# ── Style helpers ─────────────────────────────────────────────────────────────

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(hex_color, bold=False, size=10):
    return Font(color=hex_color, bold=bold, size=size, name="Calibri")

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=False)

def left():
    return Alignment(horizontal="left", vertical="center")

def hdr_border():
    orange = Side(style="medium", color=C_BORDER_HDR)
    return Border(bottom=orange)

def apply_header_style(cell, label):
    cell.value = label
    cell.fill = fill(C_HEADER_BG)
    cell.font = font(C_HEADER_TXT, bold=True, size=10)
    cell.alignment = center()
    cell.border = hdr_border()

def apply_input_style(cell, value=None, align="center"):
    if value is not None:
        cell.value = value
    cell.font = font(C_INPUT_TXT)
    cell.alignment = center() if align == "center" else left()

def apply_formula_style(cell, formula):
    cell.value = formula
    cell.font = font(C_FORMULA_TXT)
    cell.alignment = center()

def row_fill(row_idx):
    """Alternating row fill (1-indexed, skipping header row 1)."""
    return fill(C_ROW_ODD if row_idx % 2 == 1 else C_ROW_EVEN)

def write_headers(ws, headers, row=1):
    for col, hdr in enumerate(headers, 1):
        apply_header_style(ws.cell(row=row, column=col), hdr)

def set_row_bg(ws, row, num_cols):
    f = row_fill(row)
    for col in range(1, num_cols + 1):
        ws.cell(row=row, column=col).fill = f

def freeze_and_tab(ws, tab_color, freeze="B2"):
    ws.freeze_panes = freeze
    ws.sheet_properties.tabColor = tab_color

def col_width(ws, col_widths: dict):
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

def add_dropdown(ws, col_letter, start_row, end_row, formula1):
    dv = DataValidation(type="list", formula1=formula1, allow_blank=True,
                        showErrorMessage=True)
    dv.sqref = f"{col_letter}{start_row}:{col_letter}{end_row}"
    ws.add_data_validation(dv)

# ── Sheet 1: Portfolio ────────────────────────────────────────────────────────

def build_portfolio(wb):
    ws = wb.active
    ws.title = "Portfolio"

    headers = [
        "name", "arr", "arr_gr", "nav", "nav_pct", "moic", "burn",
        "gm", "nrr", "ev_arr", "ev_arr_entry", "valuation",
        "sector", "stage", "geo", "status", "funds",
        "esg_e", "esg_s", "esg_g", "lever"
    ]
    write_headers(ws, headers)
    ws.row_dimensions[1].height = 22

    # Companies: (name, arr, arr_gr, nav, moic, burn, gm, nrr,
    #              ev_arr_entry, valuation, sector, stage, geo, status,
    #              funds, esg_e, esg_s, esg_g, lever)
    # ev_arr (col J) will be =valuation/arr formula
    # nav_pct (col E) will be formula
    companies = [
        ("Q-CTRL",    18,  120, 82,  2.3, 1.2, 75, 145, 12.0,  396,  "DeepTech/Quantum", "Series D",  "AU/US", "Star",        "Fund V",  72, 80, 68, 3.2),
        ("Multiverse", 48,  65, 55,  2.0, 1.4, 66, 122,  6.2,  408,  "EdTech",           "Series D",  "UK",    "Core",        "Fund IV", 68, 71, 66, 2.8),
        ("d-Matrix",    5, 200, 38,  1.7, 4.2, 45, None, 22.0,  220,  "AI Hardware",      "Series B",  "US",    "High Growth",  "Fund V",  58, 62, 55, 4.1),
        ("Ledger",     85,  35, 32,  1.6, 0.6, 72, 118,  9.5, 1530,  "FinTech/Crypto",   "Series D",  "FR",    "Core",        "Fund III",70, 74, 69, 2.1),
        ("Wayflyer",   42,  55, 40,  1.4, 1.1, 68, 115,  5.8,  298,  "FinTech",          "Series C",  "IE/US", "Core",        "Fund V",  64, 68, 62, 2.9),
        ("Matillion",  62,  40, 32,  1.5, 1.3, 70, 130,  7.2,  583,  "Software",         "Series E",  "UK/US", "Core",        "Fund IV", 66, 70, 65, 2.5),
        ("Sanity",     24,  38, 18,  1.5, 1.8, 71, 108,  7.5,  228,  "Software",         "Series C",  "NO",    "Watch",       "Fund V",  60, 64, 58, 3.0),
        ("Headway",    15,  80, 14,  1.1, 2.1, 62,  95,  5.5,  108,  "EdTech",           "Series B",  "UA/US", "Watch",       "Fund VI", 55, 60, 52, 3.5),
        ("Booksy",     38,  42, 32,  1.3, 0.7, 71, 115,  6.8,  334,  "Software",         "Series D",  "PL/US", "Core",        "Fund V",  62, 66, 61, 2.7),
        ("Quantexa",   55,  68, 58,  2.1, 1.5, 68, 125,  9.8,  781,  "AI/Data",          "Series E",  "UK",    "Star",        "Fund IV", 71, 75, 70, 2.6),
    ]

    # Data rows 2..11
    NUM_ROWS = len(companies)
    NUM_COLS = len(headers)

    for i, co in enumerate(companies):
        r = i + 2
        (name, arr, arr_gr, nav, moic, burn, gm, nrr,
         ev_arr_entry, valuation, sector, stage, geo, status,
         funds, esg_e, esg_s, esg_g, lever) = co

        set_row_bg(ws, r, NUM_COLS)

        # Input cells
        for col, val in zip([1,2,3,4,6,7,8,9,11,12,13,14,15,16,17,18,19,20,21],
                            [name, arr, arr_gr, nav, moic, burn, gm, nrr,
                             ev_arr_entry, valuation, sector, stage, geo,
                             status, funds, esg_e, esg_s, esg_g, lever]):
            c = ws.cell(row=r, column=col)
            if val is None:
                c.value = None
            else:
                apply_input_style(c, val, align="left" if col in (1,13,14,15,16,17) else "center")

        # Formula cells: nav_pct (col 5), ev_arr (col 10)
        apply_formula_style(ws.cell(row=r, column=5),
                            f"=IFERROR(D{r}/SUM($D$2:$D$11)*100,0)")
        apply_formula_style(ws.cell(row=r, column=10),
                            f"=IFERROR(L{r}/B{r},0)")

    # Summary row (row 13)
    SUMR = NUM_ROWS + 3  # row 13
    set_row_bg(ws, SUMR, NUM_COLS)
    apply_header_style(ws.cell(row=SUMR, column=1), "TOTAL / AVG")

    num_sums  = [2,3,4,5,6,7,8,9,10]  # arr..ev_arr
    num_avgs  = [6,7,8,9]             # moic,burn,gm,nrr

    data_range = f"2:{NUM_ROWS+1}"
    for col in range(2, NUM_COLS + 1):
        c = ws.cell(row=SUMR, column=col)
        cl = get_column_letter(col)
        if col in [2, 4, 12]:   # arr, nav, valuation -> SUM
            apply_formula_style(c, f"=SUM({cl}2:{cl}{NUM_ROWS+1})")
        elif col in [3,8,9,10,18,19,20]:  # arr_gr, gm, nrr, ev_arr, esg -> AVERAGE
            apply_formula_style(c, f"=IFERROR(AVERAGE({cl}2:{cl}{NUM_ROWS+1}),0)")
        elif col in [5,6,7,11,13,14,15,16,17,21]:
            apply_formula_style(c, f"=IFERROR(AVERAGE({cl}2:{cl}{NUM_ROWS+1}),0)")
        else:
            c.value = None

    # Data validation dropdowns
    add_dropdown(ws, "P", 2, 11, '"Star,Core,High Growth,Watch"')
    add_dropdown(ws, "N", 2, 11, '"Series A,Series B,Series C,Series D,Series E,Pre-IPO"')

    # Column widths
    ws.column_dimensions["A"].width = 14
    for c in ["B","C","D","E","F","G","H","I","J","K","L"]:
        ws.column_dimensions[c].width = 9
    ws.column_dimensions["M"].width = 16
    ws.column_dimensions["N"].width = 10
    ws.column_dimensions["O"].width = 8
    ws.column_dimensions["P"].width = 12
    ws.column_dimensions["Q"].width = 9
    for c in ["R","S","T","U"]:
        ws.column_dimensions[c].width = 8

    freeze_and_tab(ws, TAB_COLORS[0])

# ── Sheet 2: Funds ────────────────────────────────────────────────────────────

def build_funds(wb):
    ws = wb.create_sheet("Funds")

    # Main fund table
    main_hdrs = ["fund", "vintage", "size_m", "irr", "tvpi", "dpi", "nav_m", "realized_m"]
    write_headers(ws, main_hdrs, row=1)

    funds_data = [
        ("Fund I",  2008, "€30M",  22.1, 3.1, 3.1,   0,  93),
        ("Fund II", 2011, "€80M",  19.8, 2.7, 2.7,   0, 200),
        ("Fund III",2013, "€150M", 17.2, 2.4, 1.9,  75, 285),
        ("Fund IV", 2018, "€220M", 16.5, 2.1, 0.8, 256, 176),
        ("Fund V",  2020, "€353M", 14.8, 1.7, 0.3, 420, 106),
        ("Fund VI", 2022, "$431M", 12.4, 1.4, 0.0, 294,   0),
    ]

    for i, row_data in enumerate(funds_data):
        r = i + 2
        set_row_bg(ws, r, len(main_hdrs))
        for col, val in enumerate(row_data, 1):
            c = ws.cell(row=r, column=col)
            apply_input_style(c, val, align="left" if col == 1 else "center")

    # Quarterly IRR trajectory section
    traj_start_row = 10
    ws.cell(row=traj_start_row, column=1).value = "Quarterly IRR Trajectory (Q2'22 – Q1'25)"
    ws.cell(row=traj_start_row, column=1).font = font(C_HEADER_TXT, bold=True)
    ws.cell(row=traj_start_row, column=1).fill = fill(C_HEADER_BG)

    quarters = ["Q2'22","Q3'22","Q4'22","Q1'23","Q2'23","Q3'23",
                "Q4'23","Q1'24","Q2'24","Q3'24","Q4'24","Q1'25"]
    traj_hdrs = ["fund"] + quarters
    write_headers(ws, traj_hdrs, row=traj_start_row + 1)

    traj_data = [
        ("Fund III", [21.2,20.8,20.1,19.5,19.0,18.6,18.2,17.8,17.5,17.3,17.2,17.2]),
        ("Fund IV",  [12.0,13.2,14.1,15.0,15.8,16.2,16.5,16.5,16.4,16.5,16.5,16.5]),
        ("Fund V",   [ 5.1, 6.8, 8.2, 9.5,10.8,12.0,12.9,13.6,14.2,14.6,14.8,14.8]),
        ("Fund VI",  [ 0.0, 0.0, 2.1, 4.5, 6.2, 7.8, 8.9,10.2,10.8,11.2,11.8,12.4]),
    ]

    for i, (fname, vals) in enumerate(traj_data):
        r = traj_start_row + 2 + i
        set_row_bg(ws, r, len(traj_hdrs))
        apply_input_style(ws.cell(row=r, column=1), fname, align="left")
        for j, v in enumerate(vals):
            apply_input_style(ws.cell(row=r, column=j + 2), v)

    # PME section
    pme_start = traj_start_row + 2 + len(traj_data) + 2
    ws.cell(row=pme_start, column=1).value = "PME Analysis"
    ws.cell(row=pme_start, column=1).font = font(C_HEADER_TXT, bold=True)
    ws.cell(row=pme_start, column=1).fill = fill(C_HEADER_BG)

    pme_hdrs = ["benchmark", "Fund I", "Fund II", "Fund III", "Fund IV", "Fund V", "Fund VI"]
    write_headers(ws, pme_hdrs, row=pme_start + 1)

    pme_data = [
        ("vs MSCI World", [1.62, 1.47, 1.38, 1.29, 1.18, 1.12]),
        ("vs NASDAQ",     [1.38, 1.25, 1.18, 1.08, 0.98, 1.02]),
    ]
    for i, (bname, vals) in enumerate(pme_data):
        r = pme_start + 2 + i
        set_row_bg(ws, r, len(pme_hdrs))
        apply_input_style(ws.cell(row=r, column=1), bname, align="left")
        for j, v in enumerate(vals):
            apply_input_style(ws.cell(row=r, column=j + 2), v)

    ws.column_dimensions["A"].width = 16
    for c in ["B","C","D","E","F","G","H","I","J","K","L","M"]:
        ws.column_dimensions[c].width = 10

    freeze_and_tab(ws, TAB_COLORS[1])

# ── Sheet 3: LPs ─────────────────────────────────────────────────────────────

def build_lps(wb):
    ws = wb.create_sheet("LPs")

    headers = ["name", "city", "type", "commit_m", "called_m", "dist_m",
               "nav_m", "irr", "tvpi", "unfunded_m"]
    write_headers(ws, headers, row=1)

    lps = [
        ("Schroder Adveq",     "Zurich",     "Fund of Funds", 65, 44.2,  8.8, 52,   14.1, 1.38),
        ("APG",                "Amsterdam",  "Pension",        50, 34.0,  6.2, 41,   13.8, 1.39),
        ("Neuberger Berman",   "New York",   "FoF",           45, 30.6,  5.4, 36,   13.5, 1.35),
        ("British Business Bk","Sheffield",  "Gov",           40, 27.2,  4.8, 32,   12.9, 1.35),
        ("EIF",                "Luxembourg", "DFI",           35, 23.8,  3.9, 27,   13.2, 1.30),
        ("Adams Street",       "Chicago",    "FoF",           30, 20.4,  3.2, 24,   12.8, 1.33),
        ("Isomer Capital",     "London",     "FoF",           25, 17.0,  2.5, 19.5, 12.6, 1.29),
        ("GP Commitment",      "London",     "GP",            20, 13.6,  2.1, 16,   14.4, 1.33),
    ]

    for i, lp in enumerate(lps):
        r = i + 2
        set_row_bg(ws, r, len(headers))
        name, city, lp_type, commit, called, dist, nav, irr, tvpi = lp
        # Input cols 1-9
        for col, val in enumerate([name, city, lp_type, commit, called, dist, nav, irr, tvpi], 1):
            align = "left" if col in (1, 2, 3) else "center"
            apply_input_style(ws.cell(row=r, column=col), val, align=align)
        # unfunded formula (col 10)
        apply_formula_style(ws.cell(row=r, column=10),
                            f"=IFERROR(D{r}-E{r},0)")

    # Summary row
    SUMR = len(lps) + 3
    set_row_bg(ws, SUMR, len(headers))
    apply_header_style(ws.cell(row=SUMR, column=1), "TOTAL / AVG")
    for col in range(2, len(headers) + 1):
        cl = get_column_letter(col)
        c = ws.cell(row=SUMR, column=col)
        if col in [4, 5, 6, 7, 10]:   # monetary -> SUM
            apply_formula_style(c, f"=SUM({cl}2:{cl}{len(lps)+1})")
        elif col in [8, 9]:            # irr, tvpi -> AVERAGE
            apply_formula_style(c, f"=IFERROR(AVERAGE({cl}2:{cl}{len(lps)+1}),0)")
        else:
            c.value = None

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 13
    ws.column_dimensions["C"].width = 13
    for c in ["D","E","F","G","H","I","J"]:
        ws.column_dimensions[c].width = 11

    freeze_and_tab(ws, TAB_COLORS[2])

# ── Sheet 4: ExitPipeline ────────────────────────────────────────────────────

def build_exit_pipeline(wb):
    ws = wb.create_sheet("ExitPipeline")

    headers = ["name", "type", "advisor", "timing",
               "bear", "base", "bull",
               "bear_prob", "base_prob", "bull_prob",
               "pwav_proceeds"]
    write_headers(ws, headers, row=1)

    exits = [
        ("Q-CTRL",    "Secondary", "Lazard",       "Q3 2025", 15, 18, 22, 0.15, 0.55, 0.30),
        ("Matillion", "IPO",       "GS/MS",        "Q4 2025", 22, 32, 48, 0.20, 0.50, 0.30),
        ("Quantexa",  "Strategic", "Internal",     "Q1 2026", 32, 46, 62, 0.20, 0.50, 0.30),
        ("Ledger",    "M&A",       "Thoma Bravo",  "Q2 2026", 38, 55, 72, 0.15, 0.55, 0.30),
        ("Wayflyer",  "IPO",       "TBD",          "2027",    22, 36, 52, 0.25, 0.45, 0.30),
    ]

    for i, ex in enumerate(exits):
        r = i + 2
        set_row_bg(ws, r, len(headers))
        name, ex_type, advisor, timing, bear, base, bull, bp, bap, bulp = ex
        for col, val in enumerate([name, ex_type, advisor, timing, bear, base, bull, bp, bap, bulp], 1):
            align = "left" if col in (1, 2, 3, 4) else "center"
            apply_input_style(ws.cell(row=r, column=col), val, align=align)
        # pwav_proceeds formula (col 11): =E*H + F*I + G*J
        apply_formula_style(ws.cell(row=r, column=11),
                            f"=IFERROR(E{r}*H{r}+F{r}*I{r}+G{r}*J{r},0)")

    # Summary row
    SUMR = len(exits) + 3
    set_row_bg(ws, SUMR, len(headers))
    apply_header_style(ws.cell(row=SUMR, column=1), "TOTAL")
    for col in [5, 6, 7, 11]:
        cl = get_column_letter(col)
        apply_formula_style(ws.cell(row=SUMR, column=col),
                            f"=SUM({cl}2:{cl}{len(exits)+1})")

    add_dropdown(ws, "B", 2, len(exits) + 1, '"Secondary,IPO,Strategic,M&A"')

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 10
    for c in ["E","F","G","H","I","J","K"]:
        ws.column_dimensions[c].width = 11

    freeze_and_tab(ws, TAB_COLORS[3])

# ── Sheet 5: Pipeline ────────────────────────────────────────────────────────

def build_pipeline(wb):
    ws = wb.create_sheet("Pipeline")

    headers = ["name", "sector", "stage", "arr_m", "arr_gr_pct", "valuation_m",
               "ev_arr", "confidence"]
    write_headers(ws, headers, row=1)

    pipeline = [
        ("Tractable",    "AI Insurance",     "Diligence", 45, 65,  540, None, "High"),
        ("Corelight",    "Cybersecurity",     "Diligence", 55, 50,  620, None, "High"),
        ("Monite",       "B2B FinTech",       "Diligence", 12, 180,  95, None, "Medium"),
        ("Rasa",         "Convo AI",          "Diligence", 18, 95,  140, None, "Medium"),
        ("Sertis",       "AI/Data",           "IC Ready",  22, 110, 175, None, "High"),
        ("Form3",        "Payments Infra",    "IC Ready",  28, 75,  225, None, "High"),
        ("Synthesia",    "AI Video",          "Tracking",  32, 85,  280, None, "Medium"),
        ("Cogni",        "Neo-banking",       "Tracking",   8, 200,  65, None, "Low"),
        ("Proxima",      "eCommerce AI",      "Tracking",  14, 90,   88, None, "Medium"),
        ("SafetyWing",   "InsurTech",         "Tracking",  22, 55,  140, None, "Medium"),
        ("Supermetrics", "Marketing Tech",    "Tracking",  42, 25,  310, None, "Low"),
        ("Tessian",      "Email Security",    "Tracking",  28, 40,  195, None, "Medium"),
    ]

    for i, co in enumerate(pipeline):
        r = i + 2
        set_row_bg(ws, r, len(headers))
        name, sector, stage, arr, arr_gr, valuation, _, confidence = co
        for col, val in enumerate([name, sector, stage, arr, arr_gr, valuation], 1):
            align = "left" if col in (1, 2, 3) else "center"
            apply_input_style(ws.cell(row=r, column=col), val, align=align)
        # ev_arr formula (col 7) = valuation / arr
        apply_formula_style(ws.cell(row=r, column=7),
                            f"=IFERROR(F{r}/D{r},0)")
        apply_input_style(ws.cell(row=r, column=8), confidence, align="center")

    add_dropdown(ws, "C", 2, len(pipeline) + 1,
                 '"Diligence,IC Ready,Tracking,Passed"')
    add_dropdown(ws, "H", 2, len(pipeline) + 1, '"High,Medium,Low"')

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 12
    for c in ["D","E","F","G","H"]:
        ws.column_dimensions[c].width = 12

    freeze_and_tab(ws, TAB_COLORS[4])

# ── Sheet 6: MacroScenarios ──────────────────────────────────────────────────

def build_macro_scenarios(wb):
    ws = wb.create_sheet("MacroScenarios")

    headers = ["scenario", "arr_gr", "gm_delta", "nrr", "ev_mult_delta",
               "dr", "timing_mo", "rates_bp", "fx_pct",
               "writeoff_pct", "probability"]
    write_headers(ws, headers, row=1)

    scenarios = [
        ("Base",           55,   0, 118,   0,  8,   0,    0,   0,  0, 30),
        ("Soft Landing",   75,   3, 128,  15,  8,  -3,  -50,   3,  0, 25),
        ("Mild Recession", 30,  -5, 105, -20, 10,   6,  100,  -5,  5, 20),
        ("Deep Recession", -10,-12,  88, -40, 13,  18,  200, -10, 15, 10),
        ("Tech Correction",35,  -2, 108, -45, 12,  12,    0,   0,  8, 10),
        ("AI Acceleration",95,   5, 138,  30,  8,  -6,    0,   5,  0,  5),
    ]

    for i, sc in enumerate(scenarios):
        r = i + 2
        set_row_bg(ws, r, len(headers))
        for col, val in enumerate(sc, 1):
            align = "left" if col == 1 else "center"
            apply_input_style(ws.cell(row=r, column=col), val, align=align)

    # Probability sum must equal 100 — add a check row
    NROWS = len(scenarios)
    sumr = NROWS + 3
    set_row_bg(ws, sumr, len(headers))
    apply_header_style(ws.cell(row=sumr, column=1), "Probability Sum")
    cl = get_column_letter(len(headers))  # col K = probability
    apply_formula_style(ws.cell(row=sumr, column=len(headers)),
                        f"=SUM({cl}2:{cl}{NROWS+1})")
    ws.cell(row=sumr, column=len(headers)).comment = None

    # Data validation: probability must be 0-100
    dv = DataValidation(type="decimal", operator="between",
                        formula1="0", formula2="100",
                        showErrorMessage=True, errorTitle="Invalid",
                        error="Probability must be between 0 and 100.")
    dv.sqref = f"K2:K{NROWS+1}"
    ws.add_data_validation(dv)

    # Note row about sum=100
    note_row = NROWS + 2
    ws.cell(row=note_row, column=1).value = "⚠ Probabilities must sum to 100"
    ws.cell(row=note_row, column=1).font = font(C_BREACH_BG, bold=True)
    ws.cell(row=note_row, column=1).fill = fill(C_HEADER_BG)

    ws.column_dimensions["A"].width = 18
    for c in ["B","C","D","E","F","G","H","I","J","K"]:
        ws.column_dimensions[c].width = 12

    freeze_and_tab(ws, TAB_COLORS[5])

# ── Sheet 7: CapitalCalls ─────────────────────────────────────────────────────

def build_capital_calls(wb):
    ws = wb.create_sheet("CapitalCalls")
    ws.cell(row=1, column=1).value = "Fund VI — Quarterly Capital Activity"
    ws.cell(row=1, column=1).font = font(C_HEADER_TXT, bold=True, size=12)
    ws.cell(row=1, column=1).fill = fill(C_HEADER_BG)

    headers = ["quarter", "calls_m", "dist_m", "nav_m",
               "cumulative_calls", "cumulative_dist", "DPI"]
    write_headers(ws, headers, row=2)
    ws.row_dimensions[2].height = 20

    quarters_data = [
        ("Q1'23",  95,  0,  90),
        ("Q2'23",  28,  0, 118),
        ("Q3'23",  22,  0, 142),
        ("Q4'23",  30,  2, 168),
        ("Q1'24",  25,  3, 198),
        ("Q2'24",  18,  4, 224),
        ("Q3'24",  22,  5, 248),
        ("Q4'24",  28,  8, 268),
        ("Q1'25",  16, 15, 294),
    ]
    NROWS = len(quarters_data)

    for i, (qtr, calls, dist, nav) in enumerate(quarters_data):
        r = i + 3  # data starts at row 3
        set_row_bg(ws, r, len(headers))
        apply_input_style(ws.cell(row=r, column=1), qtr, align="center")
        apply_input_style(ws.cell(row=r, column=2), calls)
        apply_input_style(ws.cell(row=r, column=3), dist)
        apply_input_style(ws.cell(row=r, column=4), nav)
        # cumulative_calls = running SUM from row 3 to this row
        apply_formula_style(ws.cell(row=r, column=5),
                            f"=SUM(B$3:B{r})")
        # cumulative_dist = running SUM from row 3 to this row
        apply_formula_style(ws.cell(row=r, column=6),
                            f"=SUM(C$3:C{r})")
        # DPI = cumulative_dist / 431 (Fund VI committed)
        apply_formula_style(ws.cell(row=r, column=7),
                            f"=IFERROR(F{r}/431,0)")

    # Summary row
    SUMR = NROWS + 4
    set_row_bg(ws, SUMR, len(headers))
    apply_header_style(ws.cell(row=SUMR, column=1), "TOTAL")
    for col in [2, 3, 4]:
        cl = get_column_letter(col)
        apply_formula_style(ws.cell(row=SUMR, column=col),
                            f"=SUM({cl}3:{cl}{NROWS+2})")
    # Final cumulative values
    apply_formula_style(ws.cell(row=SUMR, column=5), f"=E{NROWS+2}")
    apply_formula_style(ws.cell(row=SUMR, column=6), f"=F{NROWS+2}")
    apply_formula_style(ws.cell(row=SUMR, column=7), f"=G{NROWS+2}")

    for c in ["A","B","C","D","E","F","G"]:
        ws.column_dimensions[c].width = 16

    freeze_and_tab(ws, TAB_COLORS[6])

# ── Sheet 8: Benchmarks ──────────────────────────────────────────────────────

def build_benchmarks(wb):
    ws = wb.create_sheet("Benchmarks")

    # IRR Benchmarks
    ws.cell(row=1, column=1).value = "IRR Benchmarks by Fund"
    ws.cell(row=1, column=1).font = font(C_HEADER_TXT, bold=True, size=12)
    ws.cell(row=1, column=1).fill = fill(C_HEADER_BG)

    irr_hdrs = ["metric", "Fund I", "Fund II", "Fund III", "Fund IV", "Fund V", "Fund VI"]
    write_headers(ws, irr_hdrs, row=2)

    irr_data = [
        ("Top Quartile IRR %", [28, 25, 22, 20, 18, 16]),
        ("Median IRR %",       [22, 20, 17, 15, 13, 11]),
    ]
    for i, (mname, vals) in enumerate(irr_data):
        r = 3 + i
        set_row_bg(ws, r, len(irr_hdrs))
        apply_input_style(ws.cell(row=r, column=1), mname, align="left")
        for j, v in enumerate(vals):
            apply_input_style(ws.cell(row=r, column=j + 2), v)

    # GP fund IRR vs benchmarks
    gp_irr = [22.1, 19.8, 17.2, 16.5, 14.8, 12.4]
    r = 5
    set_row_bg(ws, r, len(irr_hdrs))
    apply_input_style(ws.cell(row=r, column=1), "GP Fund IRR %", align="left")
    for j, v in enumerate(gp_irr):
        apply_input_style(ws.cell(row=r, column=j + 2), v)

    # FX Exposure section
    fx_row = 8
    ws.cell(row=fx_row, column=1).value = "FX Exposure"
    ws.cell(row=fx_row, column=1).font = font(C_HEADER_TXT, bold=True, size=12)
    ws.cell(row=fx_row, column=1).fill = fill(C_HEADER_BG)

    fx_hdrs = ["currency", "exposure_usdm", "hedge_status", "hedge_pct"]
    write_headers(ws, fx_hdrs, row=fx_row + 1)

    fx_data = [
        ("GBP",  77, "Unhedged",      0),
        ("EUR",  32, "40% Hedged",   40),
        ("AUD",  82, "Unhedged",      0),
        ("ILS",  21, "Unhedged",      0),
    ]
    for i, (ccy, exp, status, pct) in enumerate(fx_data):
        r = fx_row + 2 + i
        set_row_bg(ws, r, len(fx_hdrs))
        apply_input_style(ws.cell(row=r, column=1), ccy, align="left")
        apply_input_style(ws.cell(row=r, column=2), exp)
        apply_input_style(ws.cell(row=r, column=3), status, align="left")
        apply_input_style(ws.cell(row=r, column=4), pct)

    # FX total
    sumr = fx_row + 2 + len(fx_data) + 1
    set_row_bg(ws, sumr, 4)
    apply_header_style(ws.cell(row=sumr, column=1), "TOTAL")
    apply_formula_style(ws.cell(row=sumr, column=2),
                        f"=SUM(B{fx_row+2}:B{fx_row+1+len(fx_data)})")

    ws.column_dimensions["A"].width = 20
    for c in ["B","C","D","E","F","G","H"]:
        ws.column_dimensions[c].width = 14

    freeze_and_tab(ws, TAB_COLORS[7])

# ── Sheet 9: WatchList ───────────────────────────────────────────────────────

def build_watchlist(wb):
    ws = wb.create_sheet("WatchList")

    headers = ["name", "flag", "threshold", "current_value", "status"]
    write_headers(ws, headers, row=1)

    # Row 2: Q-CTRL (hardcoded breach)
    r = 2
    set_row_bg(ws, r, len(headers))
    apply_input_style(ws.cell(row=r, column=1), "Q-CTRL",        align="left")
    apply_input_style(ws.cell(row=r, column=2), "Concentration", align="left")
    apply_input_style(ws.cell(row=r, column=3), "25% max",       align="center")
    apply_input_style(ws.cell(row=r, column=4), 27.9)
    # status: hardcoded BREACH with red fill
    c = ws.cell(row=r, column=5)
    c.value = "BREACH"
    c.font  = Font(color=C_WHITE, bold=True, size=10, name="Calibri")
    c.fill  = fill(C_BREACH_BG)
    c.alignment = center()

    # Row 3: d-Matrix (hardcoded WARNING)
    r = 3
    set_row_bg(ws, r, len(headers))
    apply_input_style(ws.cell(row=r, column=1), "d-Matrix",  align="left")
    apply_input_style(ws.cell(row=r, column=2), "Runway",    align="left")
    apply_input_style(ws.cell(row=r, column=3), "12mo min",  align="center")
    apply_input_style(ws.cell(row=r, column=4), "14mo")
    c = ws.cell(row=r, column=5)
    c.value = "WARNING"
    c.font  = Font(color="E8622A", bold=True, size=10, name="Calibri")
    c.fill  = fill(C_ROW_ODD)
    c.alignment = center()

    # Row 4: Sanity — VLOOKUP for arr_gr, IF formula for status
    r = 4
    set_row_bg(ws, r, len(headers))
    apply_input_style(ws.cell(row=r, column=1), "Sanity",       align="left")
    apply_input_style(ws.cell(row=r, column=2), "Growth Decel", align="left")
    apply_input_style(ws.cell(row=r, column=3), "35% min",      align="center")
    # arr_gr is column 3 in Portfolio (A=name, B=arr, C=arr_gr)
    apply_formula_style(ws.cell(row=r, column=4),
                        '=IFERROR(VLOOKUP("Sanity",Portfolio!A:C,3,0),35)')
    apply_formula_style(ws.cell(row=r, column=5),
                        '=IF(D4<35,"BREACH",IF(D4<45,"WARNING","OK"))')

    # Row 5: Headway — VLOOKUP for nrr (col 9), IF formula for status
    r = 5
    set_row_bg(ws, r, len(headers))
    apply_input_style(ws.cell(row=r, column=1), "Headway",  align="left")
    apply_input_style(ws.cell(row=r, column=2), "NRR<100%", align="left")
    apply_input_style(ws.cell(row=r, column=3), "100% min", align="center")
    # nrr is column 9 in Portfolio (I)
    apply_formula_style(ws.cell(row=r, column=4),
                        '=IFERROR(VLOOKUP("Headway",Portfolio!A:I,9,0),95)')
    apply_formula_style(ws.cell(row=r, column=5),
                        '=IF(D5<100,"BREACH","WARNING")')

    # Conditional formatting: highlight BREACH cells
    from openpyxl.formatting.rule import CellIsRule
    red_fill = PatternFill(start_color=C_BREACH_BG, end_color=C_BREACH_BG, fill_type="solid")
    ws.conditional_formatting.add(
        "E2:E5",
        CellIsRule(operator="equal", formula=['"BREACH"'], fill=red_fill,
                   font=Font(color=C_WHITE, bold=True))
    )

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 12

    freeze_and_tab(ws, TAB_COLORS[8])

# ── Sheet 10: Instructions ───────────────────────────────────────────────────

def build_instructions(wb):
    ws = wb.create_sheet("Instructions")

    lines = [
        ("GPBullhound_LiveData.xlsx — Setup & Publishing Guide", True, 14),
        ("", False, 10),
        ("1. LIVE DATA FEED — Publishing Sheet 1 (Portfolio) as CSV", True, 12),
        ("   a. Open this workbook in Excel for the web (Office 365).", False, 10),
        ('   b. Go to File > Save As > Download a Copy — or use the Share menu.', False, 10),
        ("   c. To publish as CSV: File > Export > Change File Type > CSV.", False, 10),
        ("   d. Save to OneDrive / SharePoint for continuous sync.", False, 10),
        ("   e. Copy the share link (view-only) for dashboard consumption.", False, 10),
        ("", False, 10),
        ("2. CONNECTING TO THE BULLHOUND DASHBOARD", True, 12),
        ("   a. In bullhound_dashboard.html, update DATA_CSV_URL constant", False, 10),
        ("      to the published CSV share link from step 1e.", False, 10),
        ("   b. The dashboard auto-refreshes every 5 minutes via fetch().", False, 10),
        ("   c. Ensure CORS headers are set on the CSV host if self-hosted.", False, 10),
        ("", False, 10),
        ("3. FORMULA COLOUR CONVENTION", True, 12),
        ("   Blue  (#4A8BD4) — manually entered input values.", False, 10),
        ("   Green (#2ECC8A) — formula-derived values (do NOT overwrite).", False, 10),
        ("   Red   (#E05555) — BREACH cells requiring immediate attention.", False, 10),
        ("", False, 10),
        ("4. DATA VALIDATION", True, 12),
        ("   • Portfolio!P  — status  : Star | Core | High Growth | Watch", False, 10),
        ("   • Portfolio!N  — stage   : Series A/B/C/D/E | Pre-IPO", False, 10),
        ("   • Pipeline!H   — confidence: High | Medium | Low", False, 10),
        ("   • MacroScenarios!K — probability (must sum to 100)", False, 10),
        ("", False, 10),
        ("5. RECALCULATION", True, 12),
        ("   Run:  python scripts/recalc.py", False, 10),
        ("   This script opens the workbook, iterates all cells, and reports", False, 10),
        ("   any formula errors (#DIV/0!, #REF!, #N/A, etc.).", False, 10),
        ("", False, 10),
        ("6. ADDING NEW PORTFOLIO COMPANIES", True, 12),
        ("   a. Insert a row in Portfolio rows 2–11 (extend range to 12).", False, 10),
        ("   b. Update $D$2:$D$11 references in nav_pct formulas to match.", False, 10),
        ("   c. Re-run recalc.py to confirm zero errors.", False, 10),
    ]

    for i, (text, bold, size) in enumerate(lines):
        r = i + 1
        ws.row_dimensions[r].height = 18
        c = ws.cell(row=r, column=1)
        c.value = text
        c.font  = Font(color=C_INPUT_TXT if bold else C_HEADER_TXT,
                       bold=bold, size=size, name="Calibri")
        c.fill  = fill(C_HEADER_BG if bold else C_ROW_ODD if r % 2 else C_ROW_EVEN)
        c.alignment = Alignment(horizontal="left", vertical="center")

    ws.column_dimensions["A"].width = 80
    freeze_and_tab(ws, TAB_COLORS[9])

# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    wb = Workbook()

    print("Building Sheet 1: Portfolio …")
    build_portfolio(wb)

    print("Building Sheet 2: Funds …")
    build_funds(wb)

    print("Building Sheet 3: LPs …")
    build_lps(wb)

    print("Building Sheet 4: ExitPipeline …")
    build_exit_pipeline(wb)

    print("Building Sheet 5: Pipeline …")
    build_pipeline(wb)

    print("Building Sheet 6: MacroScenarios …")
    build_macro_scenarios(wb)

    print("Building Sheet 7: CapitalCalls …")
    build_capital_calls(wb)

    print("Building Sheet 8: Benchmarks …")
    build_benchmarks(wb)

    print("Building Sheet 9: WatchList …")
    build_watchlist(wb)

    print("Building Sheet 10: Instructions …")
    build_instructions(wb)

    out_path = os.path.join(
        os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
        "GPBullhound_LiveData.xlsx"
    )
    wb.save(out_path)
    print(f"\nSaved → {out_path}")
    return out_path


if __name__ == "__main__":
    main()
